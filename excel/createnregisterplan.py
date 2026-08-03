from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill

def get_users(config: dict) -> dict[tuple[str, str], list[str]]:
    """
    Returns:
    {
        (username, password): [
            "2023-2024",
            "2024-2025",
            ...
        ]
    }
    """

    file = Path(config["input"]["create_register_plan_sheet"])

    if not file.exists():
        return {}

    workbook = load_workbook(file, data_only=True)

    if "Sheet1" not in workbook.sheetnames:
        workbook.close()
        return {}

    sheet = workbook["Sheet1"]

    headers = [
        str(cell.value).strip() if cell.value else ""
        for cell in sheet[1]
    ]

    try:
        username_col = headers.index("LOGIN ID") + 1
        password_col = headers.index("PASSWORD") + 1
    except ValueError:
        workbook.close()
        return {}

    remarks_col = None
    for i, header in enumerate(headers, start=1):
        if header.lower() == "remarks":
            remarks_col = i
            break

    # Every non-fixed column after PASSWORD until Remarks is a Financial Year
    year_cols = []

    for col in range(password_col + 1, sheet.max_column + 1):
        if remarks_col and col >= remarks_col:
            break
        year_cols.append(col)

    result = {}

    for row in range(2, sheet.max_row + 1):

        if remarks_col:
            remarks = sheet.cell(row=row, column=remarks_col).value
            if str(remarks).strip().lower() == "done":
                continue

        username = sheet.cell(row=row, column=username_col).value
        password = sheet.cell(row=row, column=password_col).value

        if not username or not password:
            continue

        years = []

        for col in year_cols:
            value = sheet.cell(row=row, column=col).value
            if value:
                years.append(str(value).strip())

        result[
            (
                str(username).strip(),
                str(password).strip(),
            )
        ] = years

    workbook.close()
    return result


def mark_users_done(
    config: dict,
    users_list: list[tuple[str, str]],
) -> int:
    """
    Marks matching users' rows green.

    Args:
        users: [(username, password), ...]

    Returns:
        Number of rows updated.
    """

    file = Path(config["input"]["create_register_plan_sheet"])

    if not file.exists():
        return 0

    workbook = load_workbook(file)

    if "Sheet1" not in workbook.sheetnames:
        workbook.close()
        return 0

    sheet = workbook["Sheet1"]

    headers = [
        str(cell.value).strip() if cell.value else ""
        for cell in sheet[1]
    ]

    try:
        username_col = headers.index("LOGIN ID") + 1
        password_col = headers.index("PASSWORD") + 1
    except ValueError:
        workbook.close()
        return 0

    # Fast lookup
    users = {
        (
            username.strip(),
            password.strip(),
        )
        for username, password in users_list
    }

    green_fill = PatternFill(
        fill_type="solid",
        start_color="92D050",
        end_color="92D050",
    )

    updated = 0

    for row in range(2, sheet.max_row + 1):
        username = str(
            sheet.cell(row=row, column=username_col).value or ""
        ).strip()

        password = str(
            sheet.cell(row=row, column=password_col).value or ""
        ).strip()

        if (username, password) in users:
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).fill = green_fill
            updated += 1

    if updated:
        workbook.save(file)

    workbook.close()

    return updated