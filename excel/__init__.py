from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook
from openpyxl.cell import Cell


USER_SHEET = "useridpass"
DATA_SHEET = "Sheet1"
STATUS_HEADER = "Status"


def get_pending_users(config) -> Dict[tuple[str, str], List[dict]]:
    """
    Returns only users that have pending records.

    Returns:
    {
        ("PR-ICHHAPADA-V-ADM", "Ichh#1234"): [
            {
                "row": 2,
                "GP": "ICHHAPADA",
                "FY": "2024-2025",
                "Activity Code": "103987549"
            },
            ...
        ],
        ...
    }
    """

    wb = load_workbook(config["input"]["excel_sheet"])

    user_ws = wb[USER_SHEET]
    data_ws = wb[DATA_SHEET]

    # ---------- Read user sheet ----------
    users = {}

    for gp, username, password in user_ws.iter_rows(
        min_row=2,
        values_only=True,
    ):
        if gp:
            users[str(gp).strip()] = {
                "username": username,
                "password": password,
            }

    # ---------- Find status column ----------
    headers = [cell.value for cell in data_ws[1]]

    status_index = None
    for i, h in enumerate(headers):
        if str(h).strip().lower() == STATUS_HEADER.lower():
            status_index = i
            break

    result = {}

    # ---------- Read pending rows ----------
    for excel_row in range(2, data_ws.max_row + 1):

        gp = data_ws.cell(excel_row, 1).value

        if gp not in users:
            continue

        status = None
        if status_index is not None:
            status = data_ws.cell(excel_row, status_index + 1).value

        if status not in (None, ""):
            continue

        key = (
            users[gp]["username"],
            users[gp]["password"],
        )

        result.setdefault(key, []).append(
            {
                "row": excel_row,
                "GP": gp,
                "FY": data_ws.cell(excel_row, 2).value,
                "AC": data_ws.cell(excel_row, 3).value,
            }
        )

    wb.close()

    return result


def update_records_status(config: dict, records: list[dict]):
    """
    Updates the status of records in Sheet1.

    If record["isDone"] is missing, it is treated as False.
    """

    path = config["input"]["excel_sheet"]

    wb = load_workbook(path)
    ws = wb[DATA_SHEET]

    headers = [cell.value for cell in ws[1]]

    if STATUS_HEADER in headers:
        status_col = headers.index(STATUS_HEADER) + 1
    else:
        status_col = ws.max_column + 1
        header = ws.cell(row=1, column=status_col)
        assert isinstance(header, Cell)
        header.value = STATUS_HEADER

    for record in records:
        row = record["row"]      # or "_row", whichever you stored
        if "isDone" not in record:
            status = ""
        elif record["isDone"] is True:
            status = "Done"
        else:
            status = "Not Done"

        cell = ws.cell(row=row, column=status_col)
        assert isinstance(cell, Cell)
        cell.value = status

    wb.save(path)
    wb.close()