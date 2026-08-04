from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.cell import Cell
from datetime import datetime,timedelta,date

def get_reconciliation_users(config: dict) -> dict[tuple[str, str], str]:
    """
    Returns:
    {
        (username, password): "bank" | "treasury" | "post_office"
    }

    Skips records whose Last Day Closed is yesterday or later.
    """

    workbook = load_workbook(
        Path(config["input"]["reconsilation_sheet"]),
        data_only=True,
    )

    try:
        records_sheet = workbook["Sheet1"]
        users_sheet = workbook["useridpass"]

        headers = [
            str(cell.value).strip() if cell.value else ""
            for cell in records_sheet[1]
        ]

        last_day_closed_col = None
        if "Last Day Closed" in headers:
            last_day_closed_col = headers.index("Last Day Closed")

        # GP -> (username, password)
        credentials: dict[str, tuple[str, str]] = {}

        for row in users_sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 3:
                continue

            gp, username, password = row[:3]

            if not gp or not username or not password:
                continue

            credentials[str(gp).strip().upper()] = (
                str(username).strip(),
                str(password).strip(),
            )

        yesterday = date.today() - timedelta(days=1)

        result: dict[tuple[str, str], str] = {}

        for row in records_sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 8:
                continue

            (
                block,
                gp,
                daybook,
                month_date,
                reconciliation,
                bank,
                treasury,
                post_office,
            ) = row[:8]

            if not gp:
                continue

            if str(reconciliation).strip().lower() != "yes":
                continue

            # Skip if Last Day Closed >= yesterday
            if last_day_closed_col is not None and last_day_closed_col < len(row):
                value = row[last_day_closed_col]

                if value:
                    closed_date = None

                    if isinstance(value, datetime):
                        closed_date = value.date()
                    elif isinstance(value, date):
                        closed_date = value
                    elif isinstance(value, str):
                        value = value.strip()

                        for fmt in (
                            "%d-%m-%Y",
                            "%d/%m/%Y",
                            "%Y-%m-%d",
                            "%d-%b-%Y",
                            "%d %b %Y",
                        ):
                            try:
                                closed_date = datetime.strptime(value, fmt).date()
                                break
                            except ValueError:
                                pass

                    if closed_date is not None and closed_date >= yesterday:
                        continue

            gp = str(gp).strip().upper()

            if gp not in credentials:
                continue

            if str(bank).strip().lower() == "yes":
                result[credentials[gp]] = "bank"
            elif str(treasury).strip().lower() == "yes":
                result[credentials[gp]] = "treasury"
            elif str(post_office).strip().lower() == "yes":
                result[credentials[gp]] = "post_office"

        return result

    finally:
        workbook.close()


def update_reconciliation_last_day_closed(
    config: dict,
    users: list[tuple[str, str, str]],
) -> bool:
    """
    Updates the 'Last Day Closed' column for the given
    (username, password, last_date_closed) tuples.

    Returns:
        True  -> workbook updated successfully
        False -> file/sheet/config error
    """

    try:
        workbook_path = Path(config["input"]["reconsilation_sheet"])
    except KeyError:
        return False

    if not workbook_path.exists():
        return False

    try:
        workbook = load_workbook(workbook_path)
    except (FileNotFoundError, InvalidFileException):
        return False

    try:
        if "Sheet1" not in workbook.sheetnames:
            return False

        if "useridpass" not in workbook.sheetnames:
            return False

        records_sheet = workbook["Sheet1"]
        users_sheet = workbook["useridpass"]

        # Find or create "Last Day Closed" column
        headers = [cell.value for cell in records_sheet[1]]

        if "Last Day Closed" in headers:
            last_day_col = headers.index("Last Day Closed") + 1
        else:
            last_day_col = len(headers) + 1
            cell = records_sheet.cell(
                row=1,
                column=last_day_col,
            )
            assert isinstance(cell,Cell)
            cell.value = "Last Day Closed"

        # (username, password) -> GP
        credentials: dict[tuple[str, str], str] = {}

        for row in users_sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 3:
                continue

            gp, username, password = row[:3]

            if not gp or not username or not password:
                continue

            credentials[
                (
                    str(username).strip(),
                    str(password).strip(),
                )
            ] = str(gp).strip().upper()

        # GP -> Last Day Closed
        gp_to_date: dict[str, str] = {}

        for username, password, last_day_closed in users:
            gp = credentials.get(
                (
                    str(username).strip(),
                    str(password).strip(),
                )
            )

            if gp:
                gp_to_date[gp] = str(last_day_closed)

        updated = False

        for row_idx, row in enumerate(records_sheet.iter_rows(min_row=2), start=2):
            gp = row[1].value

            if not gp:
                continue
            
            gp = str(gp).strip().upper()

            if gp in gp_to_date:
                cell = records_sheet.cell(
                    row=row_idx,
                    column=last_day_col,
                )
                assert isinstance(cell, Cell)
                cell.value = gp_to_date[gp]
                updated = True

        if updated:
            workbook.save(workbook_path)

        return True

    finally:
        workbook.close()